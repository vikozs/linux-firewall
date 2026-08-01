#!/usr/bin/env python3
"""
linux_firewall.py — firewalld policy audit and reconciliation for RHEL 9.

Eighth tool in the family (linux-audit, linux-harden, linux-diskspace,
linux-patch, linux-certs, linux-users, linux-drift). Reuses ssh_exec.py and
xlsx_safe.py.

You describe what each host is allowed to expose in a declarative policy file.
linux-firewall reads the effective firewalld ruleset, diffs it against the
policy, and can reconcile the difference with per-host confirmation.

    discover   read the effective ruleset, diff against policy, write a plan
               (JSON) and a report (xlsx). Changes nothing.

    apply      reconcile firewalld with the plan, RE-VALIDATING against the live
               ruleset first, one change at a time, then --reload and verify.

    report     re-render a plan to xlsx (no SSH).

Backends
--------
firewalld is the primary backend and the only one apply supports. On a host
without firewalld (or with --backend nftables) discover falls back to dumping
the nftables ruleset for visibility; it does not diff or reconcile that.

Safety
------
  * SSH lockout guard: a change that would remove the ssh service or the control
    SSH port is blocked unless --force. This is the firewall analogue of
    linux-harden's password-auth warning.
  * A host with no policy entry (and no defaults) is treated as unmanaged: its
    extra rules are reported but never proposed for removal.
  * apply re-reads the live ruleset and acts only on still-needed changes.
    Permanent changes are written, then reloaded, then verified.

Policy and plan files describe your network exposure. Keep them out of version
control (see .gitignore).
"""

import argparse
import datetime as _dt
import json
import re
import sys

from ssh_exec import SSHConfig, parse_hosts, run_fleet, run_one

__version__ = "1.0.0"
BUILD = "2026-07-31.initial"

PLAN_SCHEMA = "linux-firewall.rule-plan"
RESULT_SCHEMA = "linux-firewall.apply-result"
SCHEMA_VERSION = "1.0"

DEFAULT_SSH_PORT = "22/tcp"


# ---------------------------------------------------------------------------
# Policy loading (pure)
# ---------------------------------------------------------------------------

def load_policy(path):
    """Load a JSON or YAML policy file. YAML needs PyYAML installed."""
    with open(path) as fh:
        text = fh.read()
    if path.endswith((".yaml", ".yml")):
        try:
            import yaml
        except ImportError:
            raise ValueError("%s is YAML but PyYAML is not installed; install "
                             "PyYAML or use a JSON policy" % path)
        data = yaml.safe_load(text)
    else:
        try:
            data = json.loads(text)
        except json.JSONDecodeError as e:
            raise ValueError("%s is not valid JSON: %s" % (path, e))
    if not isinstance(data, dict):
        raise ValueError("policy must be a mapping at the top level")
    return data


def effective_policy(policy, host):
    """Merged (allowed_services, allowed_ports, managed) for one host.

    managed is False when the host has no entry and no defaults exist, so we
    never propose removing rules from a host the policy says nothing about.
    """
    groups = policy.get("groups") or {}
    hosts = policy.get("hosts") or {}
    defaults = policy.get("defaults")

    services, ports = set(), set()

    def merge(spec):
        if not spec:
            return
        services.update(spec.get("allow_services") or [])
        ports.update(_norm_port(p) for p in (spec.get("allow_ports") or []))

    merge(defaults)
    hspec = hosts.get(host) or {}
    gname = hspec.get("group")
    if gname:
        if gname not in groups:
            raise ValueError("host %s references unknown group %r" % (host, gname))
        merge(groups[gname])
    merge(hspec)

    managed = (host in hosts) or (defaults is not None)
    return services, ports, managed


def _norm_port(p):
    """Normalise '8443' -> '8443/tcp', keep '53/udp'."""
    p = str(p).strip()
    return p if "/" in p else p + "/tcp"


# ---------------------------------------------------------------------------
# Remote collector
# ---------------------------------------------------------------------------

def build_discover_script(force_nft):
    force = "1" if force_nft else "0"
    tmpl = r"""
set -u
FORCE_NFT=__FORCE_NFT__
echo "===FACTS==="
echo "hostname=$(hostname -f 2>/dev/null || hostname)"
echo "distro=$( (. /etc/os-release 2>/dev/null && echo "$PRETTY_NAME") || echo unknown)"

fw_state=absent
if command -v firewall-cmd >/dev/null 2>&1; then
  if firewall-cmd --state >/dev/null 2>&1; then fw_state=running; else fw_state=notrunning; fi
fi
echo "firewalld=$fw_state"

if [ "$fw_state" = running ] && [ "$FORCE_NFT" = 0 ]; then
  echo "backend=firewalld"
  Z="$(firewall-cmd --get-default-zone 2>/dev/null)"
  echo "default_zone=$Z"
  echo "active_zones=$(firewall-cmd --get-active-zones 2>/dev/null | awk 'NR%2==1' | tr '\n' ' ')"
  echo "===RUNTIME==="
  echo "services|$(firewall-cmd --zone="$Z" --list-services 2>/dev/null)"
  echo "ports|$(firewall-cmd --zone="$Z" --list-ports 2>/dev/null)"
  echo "===PERMANENT==="
  echo "services|$(firewall-cmd --permanent --zone="$Z" --list-services 2>/dev/null)"
  echo "ports|$(firewall-cmd --permanent --zone="$Z" --list-ports 2>/dev/null)"
else
  echo "backend=nftables"
  echo "===NFT==="
  if command -v nft >/dev/null 2>&1; then nft list ruleset 2>/dev/null; fi
fi
echo "===END==="
"""
    return tmpl.replace("__FORCE_NFT__", force)


def build_apply_script(zone, add_services, remove_services, add_ports, remove_ports):
    lines = ["set -u", 'echo "===APPLY==="', 'Z=%s' % _shquote(zone)]
    for s in add_services:
        lines.append(_fw("--add-service=%s" % s, "add service %s" % s))
    for s in remove_services:
        lines.append(_fw("--remove-service=%s" % s, "remove service %s" % s))
    for p in add_ports:
        lines.append(_fw("--add-port=%s" % p, "add port %s" % p))
    for p in remove_ports:
        lines.append(_fw("--remove-port=%s" % p, "remove port %s" % p))
    lines.append('if firewall-cmd --reload >/dev/null 2>&1; then echo "reload|ok"; else echo "reload|fail"; fi')
    # verify: emit the resulting runtime set
    lines.append('echo "verify_services|$(firewall-cmd --zone="$Z" --list-services 2>/dev/null)"')
    lines.append('echo "verify_ports|$(firewall-cmd --zone="$Z" --list-ports 2>/dev/null)"')
    lines.append('echo "===END==="')
    return "\n".join(lines) + "\n"


def _fw(op, label):
    # permanent change; report ok/fail per rule
    return ('if firewall-cmd --permanent --zone="$Z" %s >/dev/null 2>&1; '
            'then echo "%s|ok"; else echo "%s|fail"; fi' % (op, label, label))


def _shquote(s):
    if re.fullmatch(r"[A-Za-z0-9._/:-]+", s or ""):
        return s
    return "'" + str(s).replace("'", "'\\''") + "'"


# ---------------------------------------------------------------------------
# Parsing (pure)
# ---------------------------------------------------------------------------

def split_sections(stdout):
    out, cur = {}, None
    for line in stdout.splitlines():
        s = line.strip()
        m = re.match(r"^===([A-Z]+)===$", s)
        if m:
            cur = None if m.group(1) == "END" else m.group(1)
            if cur:
                out.setdefault(cur, [])
            continue
        if cur is not None:
            out[cur].append(line)
    return out


def parse_facts(lines):
    d = {}
    for line in lines:
        if "=" in line:
            k, v = line.split("=", 1)
            d[k.strip()] = v.strip()
    return d


def parse_rule_block(lines):
    """'services|a b c' / 'ports|22/tcp 80/tcp' -> (services set, ports set)."""
    services, ports = set(), set()
    for line in lines:
        if "|" not in line:
            continue
        key, val = line.split("|", 1)
        toks = set(val.split())
        if key.strip() == "services":
            services = toks
        elif key.strip() == "ports":
            ports = {_norm_port(p) for p in toks}
    return services, ports


# ---------------------------------------------------------------------------
# Diff / lockout guard (pure)
# ---------------------------------------------------------------------------

def compute_changes(allowed_s, allowed_p, runtime_s, runtime_p, ssh_port, force):
    add_s = sorted(allowed_s - runtime_s)
    add_p = sorted(allowed_p - runtime_p)
    rem_s = sorted(runtime_s - allowed_s)
    rem_p = sorted(runtime_p - allowed_p)

    blocked = []
    if not force:
        if "ssh" in rem_s:
            rem_s = [s for s in rem_s if s != "ssh"]
            blocked.append({"type": "service", "name": "ssh",
                            "reason": "would remove SSH access (use --force)"})
        if ssh_port in rem_p:
            rem_p = [p for p in rem_p if p != ssh_port]
            blocked.append({"type": "port", "name": ssh_port,
                            "reason": "would remove the control SSH port (use --force)"})
    return {"add_services": add_s, "remove_services": rem_s,
            "add_ports": add_p, "remove_ports": rem_p, "blocked": blocked}


def runtime_vs_permanent(rt_s, rt_p, pm_s, pm_p):
    out = []
    for name in sorted(rt_s ^ pm_s):
        out.append({"type": "service", "name": name,
                    "runtime": name in rt_s, "permanent": name in pm_s})
    for name in sorted(rt_p ^ pm_p):
        out.append({"type": "port", "name": name,
                    "runtime": name in rt_p, "permanent": name in pm_p})
    return out


def change_count(ch):
    return (len(ch["add_services"]) + len(ch["remove_services"])
            + len(ch["add_ports"]) + len(ch["remove_ports"]))


# ---------------------------------------------------------------------------
# Host record / plan
# ---------------------------------------------------------------------------

def host_record(res, policy, ssh_port, force):
    if not res.ok:
        return {"host": res.host, "reachable": False, "error": res.error,
                "facts": {}, "backend": None, "managed": False,
                "changes": {}, "counts": {"changes": 0, "drift_rt_perm": 0}}
    sec = split_sections(res.stdout)
    facts = parse_facts(sec.get("FACTS", []))
    backend = facts.get("backend", "unknown")
    rec = {
        "host": res.host,
        "hostname": facts.get("hostname", res.host),
        "reachable": True, "error": None, "facts": facts,
        "backend": backend, "default_zone": facts.get("default_zone", ""),
    }
    if backend != "firewalld":
        rec.update({"managed": False, "changes": {}, "unsupported": True,
                    "nft": "\n".join(sec.get("NFT", []))[:20000],
                    "counts": {"changes": 0, "drift_rt_perm": 0}})
        return rec

    rt_s, rt_p = parse_rule_block(sec.get("RUNTIME", []))
    pm_s, pm_p = parse_rule_block(sec.get("PERMANENT", []))
    allowed_s, allowed_p, managed = effective_policy(policy, res.host)
    rec["runtime"] = {"services": sorted(rt_s), "ports": sorted(rt_p)}
    rec["permanent"] = {"services": sorted(pm_s), "ports": sorted(pm_p)}
    rec["policy"] = {"services": sorted(allowed_s), "ports": sorted(allowed_p)}
    rec["managed"] = managed
    rec["rt_perm_drift"] = runtime_vs_permanent(rt_s, rt_p, pm_s, pm_p)
    if managed:
        rec["changes"] = compute_changes(allowed_s, allowed_p, rt_s, rt_p,
                                         ssh_port, force)
    else:
        rec["changes"] = {"add_services": [], "remove_services": [],
                          "add_ports": [], "remove_ports": [], "blocked": [],
                          "note": "host not in policy; reported, not reconciled"}
    rec["counts"] = {"changes": change_count(rec["changes"]),
                     "blocked": len(rec["changes"].get("blocked", [])),
                     "drift_rt_perm": len(rec["rt_perm_drift"])}
    return rec


def _now():
    return _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()


def build_plan(records, policy_path, ssh_port):
    ok = [r for r in records if r["reachable"]]
    return {
        "schema": PLAN_SCHEMA, "schema_version": SCHEMA_VERSION,
        "generated": _now(),
        "generator": {"tool": "linux-firewall", "version": __version__, "build": BUILD},
        "options": {"policy": policy_path, "ssh_port": ssh_port},
        "summary": {
            "hosts_total": len(records),
            "hosts_reachable": len(ok),
            "hosts_failed": len(records) - len(ok),
            "hosts_firewalld": sum(1 for r in ok if r["backend"] == "firewalld"),
            "hosts_unmanaged": sum(1 for r in ok if r["backend"] == "firewalld"
                                   and not r["managed"]),
            "total_changes": sum(r["counts"]["changes"] for r in ok),
            "blocked": sum(r["counts"].get("blocked", 0) for r in ok),
            "rt_perm_drift": sum(r["counts"]["drift_rt_perm"] for r in ok),
        },
        "hosts": records,
    }


def write_plan(path, plan):
    with open(path, "w") as fh:
        json.dump(plan, fh, indent=2)
        fh.write("\n")


def load_plan(path):
    try:
        with open(path) as fh:
            plan = json.load(fh)
    except json.JSONDecodeError as e:
        raise ValueError("%s is not valid JSON: %s" % (path, e))
    if not isinstance(plan, dict) or plan.get("schema") != PLAN_SCHEMA:
        raise ValueError("not a %s file: %s" % (PLAN_SCHEMA, path))
    return plan


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

NAVY = "1F3864"
HIGH = "C00000"
MED = "ED7D31"
LOW = "FFC000"
GOOD = "70AD47"

CHANGE_FILL = {
    "remove service": (HIGH, "FFFFFF"), "remove port": (HIGH, "FFFFFF"),
    "add service": (GOOD, "FFFFFF"), "add port": (GOOD, "FFFFFF"),
    "blocked": (MED, "FFFFFF"),
}


def write_report(path, plan):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from xlsx_safe import guard, safe_sheet_name, sweep, verify

    wb = Workbook()
    used = set()

    def sheet(title):
        return wb.create_sheet(safe_sheet_name(title, used))

    def header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"

    def paint(ws, col, key):
        if key in CHANGE_FILL:
            fill, font = CHANGE_FILL[key]
            cell = ws.cell(ws.max_row, col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font)

    hosts = plan["hosts"]
    ok = [h for h in hosts if h["reachable"]]

    ws = wb.active
    ws.title = safe_sheet_name("Summary", used)
    header(ws, ["Host", "Backend", "Zone", "Changes", "Blocked",
                "RT/perm drift", "Managed"])
    for h in sorted(ok, key=lambda x: -x["counts"]["changes"]):
        ws.append([h["hostname"], h.get("backend", ""), h.get("default_zone", ""),
                   h["counts"]["changes"], h["counts"].get("blocked", 0),
                   h["counts"]["drift_rt_perm"],
                   "yes" if h.get("managed") else "no"])
        for c in ws[ws.max_row]:
            guard(c)
        if h["counts"]["changes"]:
            ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=LOW)
        if h["counts"].get("blocked"):
            ws.cell(ws.max_row, 5).fill = PatternFill("solid", fgColor=MED)
            ws.cell(ws.max_row, 5).font = Font(color="FFFFFF")

    ws = sheet("Policy Drift")
    header(ws, ["Host", "Change", "Type", "Name"])
    for h in ok:
        ch = h.get("changes", {})
        rows = ([("add service", "service", s) for s in ch.get("add_services", [])]
                + [("remove service", "service", s) for s in ch.get("remove_services", [])]
                + [("add port", "port", p) for p in ch.get("add_ports", [])]
                + [("remove port", "port", p) for p in ch.get("remove_ports", [])])
        for change, typ, name in rows:
            ws.append([h["hostname"], change, typ, name])
            for c in ws[ws.max_row]:
                guard(c)
            paint(ws, 2, change)
        for b in ch.get("blocked", []):
            ws.append([h["hostname"], "blocked (%s)" % b["reason"], b["type"], b["name"]])
            for c in ws[ws.max_row]:
                guard(c)
            paint(ws, 2, "blocked")

    ws = sheet("Effective Rules")
    header(ws, ["Host", "Zone", "Runtime services", "Runtime ports",
                "Policy services", "Policy ports"])
    for h in ok:
        if h["backend"] != "firewalld":
            ws.append([h["hostname"], "", "(nftables backend, not diffed)", "", "", ""])
            for c in ws[ws.max_row]:
                guard(c)
            continue
        ws.append([h["hostname"], h.get("default_zone", ""),
                   " ".join(h["runtime"]["services"]),
                   " ".join(h["runtime"]["ports"]),
                   " ".join(h["policy"]["services"]),
                   " ".join(h["policy"]["ports"])])
        for c in ws[ws.max_row]:
            guard(c)

    ws = sheet("Runtime vs Permanent")
    header(ws, ["Host", "Type", "Name", "In runtime", "In permanent"])
    for h in ok:
        for d in h.get("rt_perm_drift", []):
            ws.append([h["hostname"], d["type"], d["name"],
                       "yes" if d["runtime"] else "", "yes" if d["permanent"] else ""])
            for c in ws[ws.max_row]:
                guard(c)
            ws.cell(ws.max_row, 3).fill = PatternFill("solid", fgColor=LOW)

    ws = sheet("Errors")
    header(ws, ["Host", "Error"])
    for h in hosts:
        if not h["reachable"]:
            ws.append([h["host"], h.get("error") or "unreachable"])
            for c in ws[ws.max_row]:
                guard(c)
                c.fill = PatternFill("solid", fgColor=HIGH)
                c.font = Font(color="FFFFFF")

    ws = sheet("About")
    s = plan["summary"]
    o = plan.get("options", {})
    about = [
        ("Tool", "linux-firewall %s" % __version__),
        ("Build", plan["generator"].get("build", BUILD)),
        ("Generated", plan["generated"]),
        ("Policy", o.get("policy", "")),
        ("Control SSH port (guarded)", o.get("ssh_port", DEFAULT_SSH_PORT)),
        ("Hosts total", s["hosts_total"]),
        ("Hosts reachable", s["hosts_reachable"]),
        ("Hosts on firewalld", s["hosts_firewalld"]),
        ("Hosts unmanaged (no policy)", s["hosts_unmanaged"]),
        ("Total changes", s["total_changes"]),
        ("Blocked by SSH guard", s["blocked"]),
        ("Runtime/permanent drift", s["rt_perm_drift"]),
        ("Note", "apply reconciles firewalld only, re-validates live, writes "
                 "permanent then reloads and verifies. Removing SSH is blocked "
                 "unless --force. Unmanaged hosts are reported, not changed."),
    ]
    for k, v in about:
        ws.append([k, v])
        guard(ws.cell(ws.max_row, 2))
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 74

    for w in wb.worksheets:
        for col in "ABCDEF":
            if w[col + "1"].value:
                w.column_dimensions[col].width = max(
                    w.column_dimensions[col].width or 0, 16)

    swept = sweep(wb)
    wb.save(path)
    bad = verify(path)
    if bad:
        raise RuntimeError("report has formula cells after sweep: %s" % bad)
    return swept


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def say(msg):
    print(msg, file=sys.stderr, flush=True)


def do_discover(hosts, cfg, policy, args):
    say("linux-firewall %s [build %s] — discover, %d host(s)"
        % (__version__, BUILD, len(hosts)))
    script = build_discover_script(args.backend == "nftables")
    records = []
    for res in run_fleet(hosts, script, cfg, workers=args.workers):
        # guard evaluated (force=False) so blocked removals surface in the plan
        rec = host_record(res, policy, args.ssh_port, force=False)
        records.append(rec)
        if not rec["reachable"]:
            say("  %-40s FAILED: %s" % (res.host, rec["error"]))
        elif rec["backend"] != "firewalld":
            say("  %-40s %s (not diffed)" % (rec["hostname"], rec["backend"]))
        else:
            say("  %-40s %d change(s)%s%s"
                % (rec["hostname"], rec["counts"]["changes"],
                   "  %d blocked" % rec["counts"]["blocked"] if rec["counts"].get("blocked") else "",
                   "  unmanaged" if not rec["managed"] else ""))
    records.sort(key=lambda r: (not r["reachable"], -r["counts"]["changes"]))
    plan = build_plan(records, args.policy, args.ssh_port)
    write_plan(args.plan, plan)
    write_report(args.output, plan)
    s = plan["summary"]
    say("\nPlan:   %s\nReport: %s" % (args.plan, args.output))
    say("Totals: %d change(s) across %d firewalld host(s), %d blocked, %d rt/perm drift"
        % (s["total_changes"], s["hosts_firewalld"], s["blocked"], s["rt_perm_drift"]))
    return plan


def _confirm(prompt, force):
    if force:
        return True
    try:
        return input(prompt + " [y/N] ").strip().lower() in ("y", "yes")
    except EOFError:
        return False


def do_apply(plan, hosts, cfg, policy, ssh_port, guard_off, confirm_skip, results_path):
    say("linux-firewall %s [build %s] — apply%s"
        % (__version__, BUILD, " (--force: SSH guard OFF)" if guard_off else ""))
    by_target = {h["target"]: h for h in hosts}
    results = []
    for prec in plan["hosts"]:
        if not prec["reachable"] or prec.get("backend") != "firewalld":
            continue
        if not prec.get("managed") or prec["counts"]["changes"] == 0:
            continue
        host = by_target.get(prec["host"]) or {"target": prec["host"],
                                               "user": None, "port": None}
        # RE-VALIDATE against live ruleset
        live = run_one(host, build_discover_script(False), cfg)
        rec = host_record(live, policy, ssh_port, force=guard_off)
        if not rec["reachable"] or rec["backend"] != "firewalld":
            results.append({"host": prec["host"], "reachable": rec["reachable"],
                            "status": "revalidation_failed"})
            say("  %-40s re-validation failed" % prec["host"])
            continue
        ch = rec["changes"]
        if change_count(ch) == 0:
            results.append({"host": prec["host"], "reachable": True,
                            "status": "already_converged", "applied": []})
            say("  %-40s already converged" % rec["hostname"])
            continue
        summary = _describe(ch)
        if ch.get("blocked"):
            say("  %-40s SSH guard blocking: %s"
                % (rec["hostname"], ", ".join(b["name"] for b in ch["blocked"])))
        if not _confirm("  Apply to %s [%s]?" % (rec["hostname"], summary), confirm_skip):
            results.append({"host": prec["host"], "reachable": True,
                            "status": "skipped"})
            say("    skipped")
            continue
        out = run_one(host, build_apply_script(
            rec["default_zone"], ch["add_services"], ch["remove_services"],
            ch["add_ports"], ch["remove_ports"]), cfg)
        parsed = _parse_apply(out)
        parsed.update({"host": prec["host"], "reachable": True,
                       "blocked": ch.get("blocked", [])})
        results.append(parsed)
        say("    %s (reload %s)" % (parsed["status"], parsed.get("reload", "?")))
    _write_results(results_path, plan, results)
    _summarize(results)
    return results


def _describe(ch):
    bits = []
    if ch["add_services"]:
        bits.append("+svc " + ",".join(ch["add_services"]))
    if ch["remove_services"]:
        bits.append("-svc " + ",".join(ch["remove_services"]))
    if ch["add_ports"]:
        bits.append("+port " + ",".join(ch["add_ports"]))
    if ch["remove_ports"]:
        bits.append("-port " + ",".join(ch["remove_ports"]))
    return "; ".join(bits) or "no changes"


def _parse_apply(res):
    if not res.ok:
        return {"status": "unreachable", "applied": [], "failed": [],
                "reload": None, "error": res.error}
    applied, failed, reload_ok = [], [], None
    for line in split_sections(res.stdout).get("APPLY", []):
        if "|" not in line:
            continue
        label, st = line.split("|", 1)
        st = st.strip()
        if label == "reload":
            reload_ok = st
        elif label.startswith(("add ", "remove ")):
            (applied if st == "ok" else failed).append(label)
    status = "applied" if not failed and reload_ok == "ok" else "partial"
    return {"status": status, "applied": applied, "failed": failed,
            "reload": reload_ok}


def _write_results(path, plan, results):
    doc = {"schema": RESULT_SCHEMA, "schema_version": SCHEMA_VERSION,
           "generated": _now(),
           "generator": {"tool": "linux-firewall", "version": __version__, "build": BUILD},
           "plan": {"generated": plan.get("generated"),
                    "generator": plan.get("generator")},
           "hosts": results}
    with open(path, "w") as fh:
        json.dump(doc, fh, indent=2)
        fh.write("\n")
    say("Results: %s" % path)


def _summarize(results):
    from collections import Counter
    c = Counter(r["status"] for r in results)
    say("\nApplied: %d  partial: %d  skipped: %d  converged: %d"
        % (c.get("applied", 0), c.get("partial", 0), c.get("skipped", 0),
           c.get("already_converged", 0)))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_cfg(args):
    cfg = SSHConfig(
        user=args.user, port=args.port, identity=args.identity,
        escalate=args.escalate, ask_ssh_pass=args.ask_ssh_pass,
        ssh_pass_env=args.ssh_pass_env, ask_sudo_pass=args.ask_sudo_pass,
        sudo_pass_same_as_ssh=args.sudo_pass_same_as_ssh,
        host_key_checking=args.host_key_checking, ssh_opts=args.ssh_opt,
        connect_timeout=args.connect_timeout, cmd_timeout=args.cmd_timeout)
    cfg.resolve_passwords()
    return cfg


def _add_conn_args(ap):
    ap.add_argument("-H", "--hosts", metavar="FILE", help="host list file")
    ap.add_argument("-u", "--user", help="default SSH user")
    ap.add_argument("-p", "--port", help="default SSH port")
    ap.add_argument("-i", "--identity", help="SSH private key")
    ap.add_argument("--escalate", choices=("none", "sudo"), default="sudo")
    ap.add_argument("--ask-ssh-pass", action="store_true")
    ap.add_argument("--ssh-pass-env", metavar="VAR")
    ap.add_argument("--ask-sudo-pass", action="store_true")
    ap.add_argument("--sudo-pass-same-as-ssh", action="store_true")
    ap.add_argument("--ssh-opt", action="append", default=[], metavar="OPT")
    ap.add_argument("--host-key-checking", default="accept-new")
    ap.add_argument("--connect-timeout", type=int, default=10)
    ap.add_argument("--cmd-timeout", type=int, default=120)
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--ssh-port", default=DEFAULT_SSH_PORT,
                    help="control SSH port the guard protects (default 22/tcp)")


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="linux-firewall",
        description="firewalld policy audit and reconciliation for a RHEL 9 fleet.")
    ap.add_argument("--version", action="version",
                    version="linux-firewall %s (build %s)" % (__version__, BUILD))
    sub = ap.add_subparsers(dest="cmd", required=True)

    d = sub.add_parser("discover", help="diff the fleet against a policy")
    _add_conn_args(d)
    d.add_argument("--policy", required=True, help="JSON or YAML policy file")
    d.add_argument("--backend", choices=("firewalld", "nftables"), default="firewalld",
                   help="nftables forces a raw ruleset dump (no diff/apply)")
    d.add_argument("--plan", default="firewall_plan.json")
    d.add_argument("-o", "--output", default="firewall_report.xlsx")

    a = sub.add_parser("apply", help="reconcile firewalld from a plan")
    _add_conn_args(a)
    a.add_argument("--plan", required=True)
    a.add_argument("--policy", required=True,
                   help="same policy used for discover (re-validated live)")
    a.add_argument("--results", default="firewall_results.json")
    a.add_argument("--force", action="store_true",
                   help="disable the SSH lockout guard (dangerous)")
    a.add_argument("-y", "--yes", dest="force_confirm", action="store_true",
                   help="skip per-host confirmation")

    rp = sub.add_parser("report", help="re-render a plan to xlsx (no SSH)")
    rp.add_argument("--plan", required=True)
    rp.add_argument("-o", "--output", default="firewall_report.xlsx")

    args = ap.parse_args(argv)

    if args.cmd == "report":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        write_report(args.output, plan)
        say("Report: %s" % args.output)
        return 0

    hosts = parse_hosts(args.hosts) if getattr(args, "hosts", None) else []
    if not hosts:
        ap.error("no hosts: pass -H/--hosts FILE")
    try:
        policy = load_policy(args.policy)
    except (ValueError, OSError) as e:
        say("error: %s" % e)
        return 2
    cfg = build_cfg(args)

    if args.cmd == "discover":
        do_discover(hosts, cfg, policy, args)
    elif args.cmd == "apply":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        ssh_port = plan.get("options", {}).get("ssh_port", args.ssh_port)
        # --force disables the SSH guard; -y/--yes skips per-host confirmation.
        do_apply(plan, hosts, cfg, policy, ssh_port,
                 guard_off=args.force, confirm_skip=args.force_confirm,
                 results_path=args.results)
    return 0


if __name__ == "__main__":
    sys.exit(main())
