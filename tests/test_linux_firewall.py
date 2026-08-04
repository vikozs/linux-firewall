import json
import os

import pytest

import linux_firewall as lf
from ssh_exec import Result

FIX = os.path.join(os.path.dirname(__file__), "fixtures")


def fixture(name):
    with open(os.path.join(FIX, name)) as fh:
        return fh.read()


def policy():
    return lf.load_policy(os.path.join(FIX, "policy.json"))


def rec(name, host, force=False):
    return lf.host_record(Result(host, ok=True, stdout=fixture(name)),
                          policy(), lf.DEFAULT_SSH_PORT, force)


# --- policy loading / effective policy --------------------------------------

def test_load_policy_json():
    p = policy()
    assert "web" in p["groups"]


def test_load_policy_rejects_non_mapping(tmp_path):
    p = tmp_path / "p.json"
    p.write_text("[1,2,3]")
    with pytest.raises(ValueError):
        lf.load_policy(str(p))


def test_effective_policy_group_merge():
    s, ports, managed = lf.effective_policy(policy(), "web01.hostname.loc")
    assert s == {"ssh", "http", "https"}
    assert ports == {"8443/tcp", "9100/tcp"}
    assert managed is True


def test_effective_policy_host_specific():
    s, ports, managed = lf.effective_policy(policy(), "db01.hostname.loc")
    assert s == {"ssh"} and ports == {"5432/tcp"} and managed


def test_effective_policy_unmanaged_host():
    s, ports, managed = lf.effective_policy(policy(), "rogue.hostname.loc")
    assert managed is False


def test_effective_policy_unknown_group_errors():
    with pytest.raises(ValueError):
        lf.effective_policy({"hosts": {"h": {"group": "nope"}}}, "h")


def test_norm_port():
    assert lf._norm_port("8443") == "8443/tcp"
    assert lf._norm_port("53/udp") == "53/udp"


# --- parsing ----------------------------------------------------------------

def test_parse_rule_block():
    s, p = lf.parse_rule_block(["services|ssh http", "ports|8443/tcp 23/tcp"])
    assert s == {"ssh", "http"} and p == {"8443/tcp", "23/tcp"}


# --- diff -------------------------------------------------------------------

def test_changes_web01():
    r = rec("discover_web01.txt", "web01.hostname.loc")
    ch = r["changes"]
    assert ch["add_services"] == ["https"]
    assert ch["remove_services"] == ["telnet"]
    assert ch["add_ports"] == ["9100/tcp"]
    assert ch["remove_ports"] == ["23/tcp"]
    assert ch["blocked"] == []
    assert r["managed"] is True
    # telnet and 23/tcp are runtime-only -> runtime/permanent drift
    drift = {(d["type"], d["name"]) for d in r["rt_perm_drift"]}
    assert ("service", "telnet") in drift and ("port", "23/tcp") in drift


def test_ssh_guard_blocks_ssh_removal():
    r = rec("discover_guard.txt", "guard.hostname.loc", force=False)
    ch = r["changes"]
    # policy omits ssh and all ports; guard must block removing ssh + 22/tcp
    assert "ssh" not in ch["remove_services"]
    assert "22/tcp" not in ch["remove_ports"]
    blocked = {(b["type"], b["name"]) for b in ch["blocked"]}
    assert ("service", "ssh") in blocked
    assert ("port", "22/tcp") in blocked


def test_force_disables_ssh_guard():
    r = rec("discover_guard.txt", "guard.hostname.loc", force=True)
    ch = r["changes"]
    assert "ssh" in ch["remove_services"]
    assert "22/tcp" in ch["remove_ports"]
    assert ch["blocked"] == []


def test_unmanaged_host_proposes_no_removals():
    r = rec("discover_unmanaged.txt", "rogue.hostname.loc")
    assert r["managed"] is False
    assert r["counts"]["changes"] == 0
    assert r["changes"].get("note")


def test_nftables_backend_not_diffed():
    r = rec("discover_nft.txt", "nftbox.hostname.loc")
    assert r["backend"] == "nftables"
    assert r.get("unsupported") is True
    assert r["counts"]["changes"] == 0
    assert "table inet filter" in r["nft"]


def test_unreachable_host():
    r = lf.host_record(Result("dead", ok=False, error="timeout"),
                       policy(), lf.DEFAULT_SSH_PORT, False)
    assert r["reachable"] is False


# --- change/count helpers ---------------------------------------------------

def test_change_count():
    ch = {"add_services": ["a"], "remove_services": ["b"],
          "add_ports": ["1/tcp"], "remove_ports": [], "blocked": []}
    assert lf.change_count(ch) == 3


def test_runtime_vs_permanent():
    d = lf.runtime_vs_permanent({"ssh", "http"}, {"80/tcp"},
                                {"ssh"}, {"80/tcp", "443/tcp"})
    names = {(x["type"], x["name"]) for x in d}
    assert ("service", "http") in names       # runtime only
    assert ("port", "443/tcp") in names        # permanent only


# --- apply script + result parsing ------------------------------------------

def test_apply_script_permanent_reload_verify():
    s = lf.build_apply_script("public", ["https"], ["telnet"],
                              ["9100/tcp"], ["23/tcp"])
    assert "--permanent" in s
    assert "--add-service=https" in s and "--remove-service=telnet" in s
    assert "--add-port=9100/tcp" in s and "--remove-port=23/tcp" in s
    assert "--reload" in s
    assert "verify_services" in s


def test_apply_script_quotes_zone():
    s = lf.build_apply_script("a;b", [], [], [], [])
    assert "'a;b'" in s


def test_parse_apply():
    body = ("===APPLY==="
            "\nadd service https|ok\nremove service telnet|ok\n"
            "add port 9100/tcp|ok\nreload|ok\n"
            "verify_services|ssh http https\n===END===\n")
    out = lf._parse_apply(Result("h", ok=True, stdout=body))
    assert out["status"] == "applied"
    assert out["reload"] == "ok"
    assert len(out["applied"]) == 3 and out["failed"] == []


def test_parse_apply_partial_on_failure():
    body = ("===APPLY===\nadd service https|fail\nreload|ok\n===END===\n")
    out = lf._parse_apply(Result("h", ok=True, stdout=body))
    assert out["status"] == "partial" and out["failed"]


# --- plan / IO --------------------------------------------------------------

def test_build_plan_summary():
    recs = [rec("discover_web01.txt", "web01.hostname.loc"),
            rec("discover_guard.txt", "guard.hostname.loc"),
            rec("discover_unmanaged.txt", "rogue.hostname.loc"),
            lf.host_record(Result("dead", ok=False, error="x"),
                           policy(), lf.DEFAULT_SSH_PORT, False)]
    plan = lf.build_plan(recs, "policy.json", lf.DEFAULT_SSH_PORT)
    s = plan["summary"]
    assert s["hosts_total"] == 4 and s["hosts_reachable"] == 3
    assert s["hosts_firewalld"] == 3
    assert s["hosts_unmanaged"] == 1
    assert s["blocked"] == 2                     # ssh service + 22/tcp on guard host
    assert plan["schema"] == lf.PLAN_SCHEMA


def test_plan_roundtrip(tmp_path):
    plan = lf.build_plan([rec("discover_web01.txt", "web01.hostname.loc")],
                         "policy.json", lf.DEFAULT_SSH_PORT)
    p = tmp_path / "plan.json"
    lf.write_plan(str(p), plan)
    assert lf.load_plan(str(p))["summary"]["total_changes"] == 4


def test_load_plan_rejects_wrong_schema(tmp_path):
    p = tmp_path / "bad.json"
    p.write_text(json.dumps({"schema": "nope"}))
    with pytest.raises(ValueError):
        lf.load_plan(str(p))


# --- report -----------------------------------------------------------------

def test_write_report_is_formula_clean(tmp_path):
    from xlsx_safe import verify
    recs = [rec("discover_web01.txt", "web01.hostname.loc"),
            rec("discover_guard.txt", "guard.hostname.loc"),
            rec("discover_nft.txt", "nftbox.hostname.loc"),
            lf.host_record(Result("dead", ok=False, error="x"),
                           policy(), lf.DEFAULT_SSH_PORT, False)]
    plan = lf.build_plan(recs, "policy.json", lf.DEFAULT_SSH_PORT)
    out = tmp_path / "r.xlsx"
    lf.write_report(str(out), plan)
    assert verify(str(out)) == {}
    from openpyxl import load_workbook
    names = load_workbook(str(out)).sheetnames
    for s in ("Summary", "Policy Drift", "Effective Rules",
              "Runtime vs Permanent", "Errors", "About"):
        assert s in names


def test_report_neutralizes_formula_injection(tmp_path):
    from xlsx_safe import verify
    body = ("===FACTS===\nhostname==cmd|'/C calc'!A1\nfirewalld=running\n"
            "backend=firewalld\ndefault_zone=public\n===RUNTIME===\n"
            "services|ssh\nports|22/tcp\n===PERMANENT===\nservices|ssh\n"
            "ports|22/tcp\n===END===\n")
    r = lf.host_record(Result("db01.hostname.loc", ok=True, stdout=body),
                       policy(), lf.DEFAULT_SSH_PORT, False)
    plan = lf.build_plan([r], "policy.json", lf.DEFAULT_SSH_PORT)
    out = tmp_path / "r.xlsx"
    lf.write_report(str(out), plan)
    assert verify(str(out)) == {}


# --- collector --------------------------------------------------------------

def test_discover_scripts_valid_bash():
    import subprocess
    for s in (lf.build_discover_script(False), lf.build_discover_script(True),
              lf.build_apply_script("public", ["http"], ["telnet"],
                                    ["8443/tcp"], ["23/tcp"])):
        r = subprocess.run(["bash", "-n"], input=s, text=True,
                           capture_output=True)
        assert r.returncode == 0, r.stderr
